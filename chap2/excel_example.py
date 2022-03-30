import openpyxl

wb = openpyxl.load_workbook('./sample.xlsx')
sheet = wb['Sheet1']
print(sheet.max_column, sheet.max_row)
print(sheet.cell(row=1, column=1).value)
print(sheet.cell(row=2, column=2).value)
print()
for row in sheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value)
    print('-' * 10)
wb.close()